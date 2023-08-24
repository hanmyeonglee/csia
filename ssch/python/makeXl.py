from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet as Sheet
from openpyxl.styles import Border, Side, Font, Alignment
from math import ceil
from copy import deepcopy as copy


def divider(data, height):
    form = {'id': '', 'number': '', 'name': '',
            'sex': '', 'disease': '', 'treat': '', 'time': ''}
    head = {'id': '연변', 'number': '학년반', 'name': '성명',
            'sex': '성별', 'disease': '병명', 'treat': '처 치', 'time': '시간'}
    if len(data) <= 46:
        data.extend([copy(form) for _ in range(46 - len(data))])
        ret01, ret02 = [copy(head)], [copy(head)]
        ret01 += (data[:23])
        ret02 += (data[23:])
        return ret01, ret02
    else:
        if len(data) % 2 == 0:
            ret01, ret02 = [copy(head)], [copy(head)]
            ret01 += (data[:height])
            ret02 += (data[height:])
            return ret01, ret02
        else:
            ret01, ret02 = [copy(head)], [copy(head)]
            ret01 += (data[:height-1])
            ret02 += (data[height-1:])
            ret02.append(copy(form))
            return ret01, ret02


def makeFile(data: list, stat: list, dateFileName: str):
    wb = Workbook()
    sheet: Sheet = wb.active
    sheet.page_margins.left = 0.2519685
    sheet.page_margins.right = 0.2519685

    sheet.column_dimensions['A'].width = 4.95
    for i in range(18):
        sheet.column_dimensions[chr(i+66)].width = 4.79

    height = 24 if len(data) <= 46 else ceil(len(data)/2)
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    header = {'A1:N3': '보   건   일   지', 'A4:N4': '2023년 4월 3일 월요일', 'O1:O4': '결\n재', 'P1:Q1': '계', 'P2:Q4': '', 'R1:S1': '부장', 'R2:S4': '', 'A5:C7': '보건교육',
              'D5:J7': '', 'K5:M7': '보건행사', 'N5:S7': '', 'A8:C10': '병원치료', 'D8:J10': '', 'K8:M10': '특이사항', 'N8:S10': ''}

    d1, d2 = divider(data, height)
    body = {f'A11:A{10+height}': '응\n\n급\n\n처\n\n치'}
    for i in range(11, 11+height):
        dic01 = d1[i-11]
        dic02 = d2[i-11]

        upd01 = {
            f'B{i}:B{i}': dic01['id'],
            f'C{i}:C{i}': dic01['number'],
            f'D{i}:D{i}': dic01['name'],
            f'E{i}:E{i}': dic01['sex'],
            f'F{i}:G{i}': dic01['disease'],
            f'H{i}:J{i}': f'{dic01["treat"]}({dic01["time"]})' if dic01['id'] != '' else ''
        }
        upd02 = {
            f'K{i}:K{i}': dic02['id'],
            f'L{i}:L{i}': dic02['number'],
            f'M{i}:M{i}': dic02['name'],
            f'N{i}:N{i}': dic02['sex'],
            f'O{i}:P{i}': dic02['disease'],
            f'Q{i}:S{i}': f'{dic02["treat"]}({dic02["time"]})' if dic02['id'] != '' else ''
        }

        body.update(upd01)
        body.update(upd02)

    footer = {f'A{11+height}:A{17+height}': '통\n\n계', f'B{11+height}:B{11+height}': '종류', f'B{12+height}:B{13+height}': '일계',
              f'B{14+height}:B{15+height}': '월계', f'B{16+height}:B{17+height}': '누계'}

    for i, row in enumerate(sheet.iter_rows(min_row=11+height, max_row=17+height, min_col=3, max_col=19)):
        for cell, value in zip(row, stat[i]):
            cell.value = value

    for key, val in header.items():
        sheet.merge_cells(key)
        sheet[key.split(':')[0]] = val
    for key, val in body.items():
        sheet.merge_cells(key)
        sheet[key.split(':')[0]] = val
    for key, val in footer.items():
        sheet.merge_cells(key)
        sheet[key.split(':')[0]] = val

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(
                wrap_text=True, horizontal='center', vertical='center')

    for row in sheet.iter_rows(min_row=12, max_row=10+height, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True, horizontal='left', vertical='center')

    for row in sheet.iter_rows(min_row=12, max_row=10+height, min_col=17, max_col=17):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True, horizontal='left', vertical='center')

    sheet['A1'].font = Font(size=24, bold=True)

    """ for row in sheet.iter_rows(min_row=11, max_row=11, min_col=2, max_col=7):
        values = ['연변', '학년반', '성명', '성별', '병명', '처\t치']
        for cell, value in zip(row, values):
            cell.value = value

    for row in sheet.iter_rows(min_row=11, max_row=11, min_col=8, max_col=13):
        values = ['연변', '학년반', '성명', '성별', '병명', '처\t치']
        for cell, value in zip(row, values):
            cell.value = value """

    wb.save(f'/var/www/html/ssch/csv/{dateFileName}.xlsx')


